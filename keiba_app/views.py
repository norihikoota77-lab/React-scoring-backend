import os
import random
import pandas as pd
import openpyxl
import tempfile
import traceback
import urllib.parse
from django.conf import settings
from django.shortcuts import render
from django.http import FileResponse, Http404
from .forms import UploadForm
from .scoring_engine import ScoringEngine
from django.http import JsonResponse #★追加
from django.views.decorators.csrf import csrf_exempt
from .models import ScoreHistory
from django.forms.models import model_to_dict
import csv
from django.http import HttpResponse
from django.utils import timezone


def index(request):
    """ファイル選択および処理のビュー"""
    form = UploadForm()
    
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        
        if form.is_valid():
            correct_file = request.FILES["correct_file"]
            user_file = request.FILES["user_file"]

            # アップロードされたファイルを一時ファイルとして保持
            c_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            u_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            
            try:
                for chunk in correct_file.chunks():
                    c_temp.write(chunk)
                for chunk in user_file.chunks():
                    u_temp.write(chunk)
                
                c_temp.close()
                u_temp.close()
                
                c_path = c_temp.name
                u_path = u_temp.name

                # 正解ファイルのB13セルからタイトルを取得
                correct_title = "タイトル不明"
                try:
                    wb_c = openpyxl.load_workbook(c_path, data_only=True)
                    ws_c = wb_c.active
                    if ws_c['B13'].value:
                        correct_title = str(ws_c['B13'].value)
                except Exception as e:
                    pass

                # ユーザーファイルのB14セルから名前を取得
                user_name = "名無し"
                try:
                    wb_u = openpyxl.load_workbook(u_path, data_only=True)
                    ws_u = wb_u.active

                    print(ws_u["B14"].value)

                    if ws_u['B14'].value:
                        raw_name = str(ws_u['B14'].value)
                        user_name = "".join(c for c in raw_name if c not in r'\/:*?"<>|')
                except Exception as e:
                    pass

                # パス文字列へ変換して結合し、保存先フォルダを作成
                output_dir = os.path.join(str(settings.BASE_DIR), "reports")
                os.makedirs(output_dir, exist_ok=True)
                
                output_file = os.path.join(
                    output_dir, f"{user_name}_{os.path.splitext(user_file.name)[0]}_レース結果報告.xlsx"
                )

                engine = ScoringEngine()
                engine.grade(c_path, u_path)
                engine.export_excel(output_file)

                msg, color = engine.get_result_message()

                # スコアに応じた動画フォルダの振り分け
                if engine.percentage >= 80:
                    folder_name = "excellent"
                elif engine.percentage >= 50:
                    folder_name = "good"
                else:
                    folder_name = "try_again"

                static_videos_dir = os.path.join(str(settings.BASE_DIR), 'keiba_app', 'static', 'videos')
                folder_path = os.path.join(static_videos_dir, folder_name)
                video_file = f"videos/{folder_name}/default.mp4"

                if os.path.exists(folder_path):
                    mp4_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.mp4')]
                    if mp4_files:
                        selected_file = random.choice(mp4_files)
                        video_file = f"videos/{folder_name}/{selected_file}"

                file_name = os.path.basename(output_file)

                report_html = ""
                if os.path.exists(output_file):
                    try:
                        df = pd.read_excel(output_file)

                        # 空列名を空文字へ
                        df.columns = [
                            "" if "Unnamed" in str(col) else str(col).split(".")[0]
                            for col in df.columns
                        ]   

                        # 空セルを空文字へ
                        df = df.fillna("")
 
                        # float → int文字列化
                        df = df.map(
                            lambda x: str(int(x))
                            if isinstance(x, float) and x.is_integer()
                            else x
                        )
 
                        # HTML装飾
                        df = df.replace("⭕", '<span class="ok">⭕</span>')
                        df = df.replace("✖", '<span class="ng">✖</span>')

                        report_html = df.to_html(
                            classes='dataframe',
                            border=0,
                            index=False,
                            justify="center",
                            escape=False
                        )
                        

                        report_html = report_html.replace('✖', '<span class="text-blue-600 font-bold">✖</span>')
                    except Exception as e:
                        report_html = f"<p class='text-red-500'>プレビュー読み込みエラー: {e}</p>"

                context = {
                    "score": engine.score,
                    "valid_count": engine.valid_count,
                    "percentage": engine.percentage,
                    "rank": engine.get_rank(),
                    "msg": msg,
                    "color": color,
                    "report_file_name": file_name,
                    "report_html": report_html,
                    "video_file": video_file,
                    "user_name": user_name,
                    "correct_title": correct_title, 
                    "rows_data": engine.rows_data,
 
                        # 追加
                    "table_sets": [
                          engine.rows_data[:20],
                          engine.rows_data[20:40],
                    ],
                }

                return render(request, "result.html", context)

            except Exception as e:
                print("--- [採点処理でエラーが発生しました] ---")
                traceback.print_exc()
                return render(request, "index.html", {"form": form, "error": f"処理中にエラーが発生しました: {str(e)}"})

            finally:
                # 一時ファイルの確実に削除
                if os.path.exists(c_temp.name):
                    os.remove(c_temp.name)
                if os.path.exists(u_temp.name):
                    os.remove(u_temp.name)
        
        else:
            return render(request, "index.html", {"form": form, "error": "入力されたファイルが無効です。"})

    return render(request, "index.html", {"form": form})


def download_report(request, file_name):
    """ファイルをダウンロードさせる専用ビュー"""
    decoded_file_name = urllib.parse.unquote(file_name)
    file_path = os.path.join(str(settings.BASE_DIR), 'reports', decoded_file_name)
    
    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'), as_attachment=True)
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urllib.parse.quote(decoded_file_name)}"
        return response
    else:
        raise Http404(f"ファイルが見つかりません: {file_path}")
    

#追加
@csrf_exempt
def score_api(request):

    if request.method != "POST":
        return JsonResponse(
            {"error": "POST only"},
            status=400
        )

    try:

        correct_file = request.FILES["correct_file"]
        user_file = request.FILES["user_file"]

        c_temp = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        )

        u_temp = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        )

        for chunk in correct_file.chunks():
            c_temp.write(chunk)

        for chunk in user_file.chunks():
            u_temp.write(chunk)

        c_temp.close()
        u_temp.close()

        engine = ScoringEngine()

        user_name = "名無し"

        exam_title = "問題"

        try:

            wb_c = openpyxl.load_workbook(
                c_temp.name,
                data_only=True
            )

            ws_c = wb_c.active

            if ws_c["B13"].value:

                exam_title = str(
                    ws_c["B13"].value
                )

        except:
            pass


        try:

            wb_u = openpyxl.load_workbook(
                u_temp.name,
                data_only=True
            )

            ws_u = wb_u.active

            if ws_u["B14"].value:

                user_name = str(
                    ws_u["B14"].value
                )

        except:
            pass

    
        try:
            engine.grade(
                c_temp.name,
                u_temp.name
            )
        except ValueError as e:
            # 試験名不一致の場合はここで返す
            return JsonResponse(
                {"error": str(e)},
                status=400
            )

        msg, color = engine.get_result_message()

        ScoreHistory.objects.create(
            score=engine.score,
            valid_count=engine.valid_count,
            percentage=engine.percentage,
            rank=engine.get_rank(),
            message=msg,
            rows_data=engine.rows_data,
            user_name=user_name,
            exam_title=exam_title,
        ) 

        # スコアに応じた動画フォルダ
        if engine.percentage >= 80:
            folder_name = "excellent"
        elif engine.percentage >= 50:
            folder_name = "good"
        else:
            folder_name = "try_again"

        static_videos_dir = os.path.join(
            str(settings.BASE_DIR),
            "keiba_app",
            "static",
            "videos"
        )

        folder_path = os.path.join(
            static_videos_dir,
            folder_name
        )

        video_file = f"videos/{folder_name}/default.mp4"

        if os.path.exists(folder_path):

            mp4_files = [
                f for f in os.listdir(folder_path)
                if f.lower().endswith(".mp4")
            ]

            if mp4_files:

                selected_file = random.choice(mp4_files)

                video_file = (
                    f"videos/{folder_name}/{selected_file}"
                )

        return JsonResponse({

            "score": engine.score,
            "valid_count": engine.valid_count,
            "percentage": engine.percentage,
            "rank": engine.get_rank(),
            "msg": msg,
            "rows_data": engine.rows_data,
            "video_file": video_file,       
            "user_name": user_name,
            "exam_title": exam_title,

        })

    except Exception as e:

        return JsonResponse(
            {"error": str(e)},
            status=500
        )

    finally:

        if os.path.exists(c_temp.name):
            os.remove(c_temp.name)

        if os.path.exists(u_temp.name):
            os.remove(u_temp.name)

@csrf_exempt
def history_api(request):

    histories = ScoreHistory.objects.order_by(
        "-created_at"
    )[:20]

    data = []

    for history in histories:

        data.append({
            "id": history.id,
            "score": history.score,
            "valid_count": history.valid_count,
            "percentage": history.percentage,
            "rank": history.rank,
            "message": history.message,
            "rows_data": history.rows_data,
            "created_at": history.created_at.strftime("%Y-%m-%d %H:%M"),
            "user_name": history.user_name,
            "exam_title": history.exam_title,
            "created_at": timezone.localtime(
                history.created_at
            ).strftime("%Y-%m-%d %H:%M"),

        })

    return JsonResponse(data, safe=False)


@csrf_exempt
def delete_history_api(request, history_id):

    if request.method != "DELETE":

        return JsonResponse(
            {"error": "DELETE only"},
            status=400
        )

    try:

        history = ScoreHistory.objects.get(
            id=history_id
        )

        history.delete()

        return JsonResponse({

            "message": "deleted"

        })

    except ScoreHistory.DoesNotExist:

        return JsonResponse(
            {"error": "not found"},
            status=404
        )


def export_history_csv(request):

    histories = ScoreHistory.objects.all().order_by(
        "-created_at"
    )

    response = HttpResponse(
        content_type="text/csv; charset=utf-8-sig"
    )

    response.write('\ufeff')

    response[
        "Content-Disposition"
    ] = 'attachment; filename="score_history.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "日時",
        "スコア",
        "正答率",
        "ランク",
        "メッセージ",
    ])

    for history in histories:

        writer.writerow([

            timezone.localtime(
                history.created_at
            ).strftime("%Y-%m-%d %H:%M:%S"),

            history.score,

            history.percentage,

            history.rank,

            history.message,

        ])

    return response

# ================================================
#  Web採点 API
# ================================================

from .models import Exam, Question


@csrf_exempt
def exam_list_api(request):
    """試験一覧を返す"""

    if request.method != "GET":
        return JsonResponse({"error": "GET only"}, status=400)

    exams = Exam.objects.all().order_by("title")

    data = [
        {
            "id": exam.id,
            "title": exam.title,
            "choice_type": exam.choice_type,
            "question_count": exam.questions.count(),
            "show_questions": exam.show_questions,  # ★追加            
        }
        for exam in exams
    ]

    return JsonResponse(data, safe=False)


@csrf_exempt
def exam_questions_api(request, exam_id):
    """指定試験の問題一覧を返す"""

    if request.method != "GET":
        return JsonResponse({"error": "GET only"}, status=400)

    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return JsonResponse({"error": "試験が見つかりません"}, status=404)

    questions = exam.questions.all()

    data = {
        "exam_id": exam.id,
        "exam_title": exam.title,
        "choice_type": exam.choice_type,
        "questions": [
            {
                "number": q.number,
                "text": q.text,
            }
            for q in questions
        ],
    }

    return JsonResponse(data)


@csrf_exempt
def exam_submit_api(request, exam_id):
    """解答を受け取って採点し結果を返す"""

    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return JsonResponse({"error": "試験が見つかりません"}, status=404)

    try:
        import json
        body = json.loads(request.body)
        user_name = body.get("user_name", "名無し")
        answers = body.get("answers", {})
        # answers = {"1": "3", "2": "A", ...} 問題番号: 解答

        questions = exam.questions.all()

        rows_data = []
        score = 0
        valid_count = 0

        for q in questions:
            user_ans = str(answers.get(str(q.number), "")).strip().upper()
            correct_ans = str(q.correct_answer).strip().upper()

            is_correct = user_ans == correct_ans
            valid_count += 1

            if is_correct:
                score += 1

            rows_data.append([
                q.number,
                user_ans if user_ans else "未記入",
                correct_ans,
                "⭕" if is_correct else "✖",
            ])

        percentage = (score / valid_count * 100) if valid_count > 0 else 0.0

        # ランク判定
        if percentage == 100:
            rank = "S"
        elif percentage >= 70:
            rank = "A"
        elif percentage >= 50:
            rank = "B"
        else:
            rank = "C"

        RESULT_MESSAGES = {
            "S": "🌟🏆 [G1制覇] 伝説の三冠馬級！",
            "A": "🥈 [重賞入着] 素晴らしい末脚です",
            "B": "🐎 [入賞] 掲示板に載りました",
            "C": "🏃 [未勝利] ゲート練習からやり直し",
        }

        msg = RESULT_MESSAGES[rank]

        # 動画選択
        if percentage >= 80:
            folder_name = "excellent"
        elif percentage >= 50:
            folder_name = "good"
        else:
            folder_name = "try_again"

        static_videos_dir = os.path.join(
            str(settings.BASE_DIR),
            "keiba_app",
            "static",
            "videos"
        )

        folder_path = os.path.join(static_videos_dir, folder_name)
        video_file = f"videos/{folder_name}/default.mp4"

        if os.path.exists(folder_path):
            mp4_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".mp4")]
            if mp4_files:
                video_file = f"videos/{folder_name}/{random.choice(mp4_files)}"

        # 履歴保存
        ScoreHistory.objects.create(
            score=score,
            valid_count=valid_count,
            percentage=percentage,
            rank=rank,
            message=msg,
            rows_data=rows_data,
            user_name=user_name,
            exam_title=exam.title,
        )

        return JsonResponse({
            "score": score,
            "valid_count": valid_count,
            "percentage": percentage,
            "rank": rank,
            "msg": msg,
            "rows_data": rows_data,
            "video_file": video_file,
            "user_name": user_name,
            "exam_title": exam.title,
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    

# ================================================
#  Web解答採点 API
# ================================================

@csrf_exempt
def score_web_api(request):
    """Web解答データと正解マスタExcelを受け取って採点する"""

    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    try:
        correct_file = request.FILES["correct_file"]
        user_name = request.POST.get("user_name", "名無し")
        exam_title = request.POST.get("exam_title", "試験")
        answers_json = request.POST.get("answers", "{}")
        question_count = int(request.POST.get("question_count", 40))

        import json
        answers = json.loads(answers_json)
        # answers = {"1": "3", "2": "A", ...}

        # 正解マスタを一時ファイルに保存
        c_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        for chunk in correct_file.chunks():
            c_temp.write(chunk)
        c_temp.close()

        # 試験名チェック（正解マスタのB13と入力された試験名を比較）
        try:
            wb_c = openpyxl.load_workbook(c_temp.name, data_only=True)
            ws_c = wb_c.active
            if ws_c["B13"].value:
                correct_exam_title = str(ws_c["B13"].value).strip()
                if correct_exam_title != exam_title.strip():
                    return JsonResponse(
                        {
                            "error": f"試験名が一致しません\n正解マスタ: {correct_exam_title}\n入力値: {exam_title}"
                        },
                        status=400,
                    )
        except Exception:
            pass

        # 正解マスタを読み込む
        engine = ScoringEngine()
        correct_map = engine.load_answers(c_temp.name)

        # 採点
        rows_data = []
        score = 0
        valid_count = 0

        for q_num in range(1, question_count + 1):
            correct_ans = correct_map.get(q_num)
            user_ans = str(answers.get(str(q_num), "")).strip().upper()

            is_valid = correct_ans is not None
            is_correct = user_ans == str(correct_ans).strip().upper() if is_valid else False

            if is_valid:
                valid_count += 1
                if is_correct:
                    score += 1

            rows_data.append([
                q_num,
                user_ans if user_ans else "未記入",
                correct_ans if is_valid else "-",
                "⭕" if (is_valid and is_correct) else ("✖" if is_valid else "-"),
            ])

        percentage = (score / valid_count * 100) if valid_count > 0 else 0.0

        # ランク判定
        if percentage == 100:
            rank = "S"
        elif percentage >= 70:
            rank = "A"
        elif percentage >= 50:
            rank = "B"
        else:
            rank = "C"

        RESULT_MESSAGES = {
            "S": "🌟🏆 [G1制覇] 伝説の三冠馬級！",
            "A": "🥈 [重賞入着] 素晴らしい末脚です",
            "B": "🐎 [入賞] 掲示板に載りました",
            "C": "🏃 [未勝利] ゲート練習からやり直し",
        }
        msg = RESULT_MESSAGES[rank]

        # 動画選択
        if percentage >= 80:
            folder_name = "excellent"
        elif percentage >= 50:
            folder_name = "good"
        else:
            folder_name = "try_again"

        static_videos_dir = os.path.join(str(settings.BASE_DIR), "keiba_app", "static", "videos")
        folder_path = os.path.join(static_videos_dir, folder_name)
        video_file = f"videos/{folder_name}/default.mp4"

        if os.path.exists(folder_path):
            mp4_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".mp4")]
            if mp4_files:
                video_file = f"videos/{folder_name}/{random.choice(mp4_files)}"

        # 履歴保存
        ScoreHistory.objects.create(
            score=score,
            valid_count=valid_count,
            percentage=percentage,
            rank=rank,
            message=msg,
            rows_data=rows_data,
            user_name=user_name,
            exam_title=exam_title,
        )

        return JsonResponse({
            "score": score,
            "valid_count": valid_count,
            "percentage": percentage,
            "rank": rank,
            "msg": msg,
            "rows_data": rows_data,
            "video_file": video_file,
            "user_name": user_name,
            "exam_title": exam_title,
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    finally:
        if os.path.exists(c_temp.name):
            os.remove(c_temp.name)



@csrf_exempt
def create_superuser(request):
    from django.contrib.auth.models import User
    User.objects.filter(username="admin").delete()
    user = User.objects.create(username="admin", is_superuser=True, is_staff=True)
    user.set_password("Admin1234!")
    user.save()
    return JsonResponse({"message": "再作成完了"})

# ================================================
#  Excel詳細レポートダウンロード API
# ================================================

@csrf_exempt
def download_excel_api(request):
    """採点データを受け取ってExcelレポートを返す"""

    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    try:
        import json
        from django.http import FileResponse
        import io

        body = json.loads(request.body)
        user_name = body.get("user_name", "名無し")
        exam_title = body.get("exam_title", "試験")
        score = body.get("score", 0)
        valid_count = body.get("valid_count", 0)
        percentage = body.get("percentage", 0)
        rank = body.get("rank", "C")
        pass_score = body.get("pass_score", 70)
        rows_data = body.get("rows_data", [])

        # 合格・不合格判定
        is_pass = float(percentage) >= float(pass_score)
        pass_label = "合格" if is_pass else "不合格"

        # pandas でExcel作成
        mid = len(rows_data) // 2
        left_rows = rows_data[:mid] if mid > 0 else rows_data
        right_rows = rows_data[mid:] if mid > 0 else []

        left_df = pd.DataFrame(
            left_rows,
            columns=["問題", "解答", "正解", "判定"]
        )
        right_df = pd.DataFrame(
            right_rows,
            columns=["問題", "解答", "正解", "判定"]
        )

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            left_df.to_excel(
                writer,
                sheet_name="採点結果",
                startrow=7,
                startcol=0,
                index=False
            )
            if right_rows:
                right_df.to_excel(
                    writer,
                    sheet_name="採点結果",
                    startrow=7,
                    startcol=5,
                    index=False
                )

            ws = writer.sheets["採点結果"]

            # ヘッダー情報
            ws["A1"] = "🐎 競馬演出スコアラー"
            ws["A2"] = f"試験名：{exam_title}"
            ws["A3"] = f"受験者：{user_name}"
            ws["A5"] = f"スコア：{score} / {valid_count}"
            ws["A6"] = f"正答率：{round(float(percentage), 1)}%"
            ws["D5"] = f"ランク：{rank}"
            ws["D6"] = f"判定：{pass_label}"

            # スタイル設定
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            # タイトルスタイル
            ws["A1"].font = Font(bold=True, size=16)

            # 判定色
            pass_color = "1B5E20" if is_pass else "B71C1C"
            ws["D6"].font = Font(bold=True, color=pass_color, size=12)

            # テーブルスタイル
            header_fill = PatternFill(start_color="DA1F28", end_color="DA1F28", fill_type="solid")
            ok_fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
            ng_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            START_ROW = 8
            target_cols_left = [1, 2, 3, 4]
            target_cols_right = [6, 7, 8, 9]

            total_rows = max(len(left_rows), len(right_rows))

            for r in range(START_ROW, START_ROW + total_rows + 1):
                for col_group in [target_cols_left, target_cols_right]:
                    for c in col_group:
                        cell = ws.cell(row=r, column=c)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                        if r == START_ROW:
                            cell.fill = header_fill
                            cell.font = Font(color="FFFFFF", bold=True)
                        else:
                            row_idx = r - START_ROW - 1
                            if col_group == target_cols_left and row_idx < len(left_rows):
                                judgment = left_rows[row_idx][3]
                                if c in [2]:
                                    cell.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
                                elif judgment == "⭕":
                                    cell.fill = ok_fill
                                elif judgment == "✖":
                                    cell.fill = ng_fill
                                if c == 4:
                                    cell.font = Font(
                                        color="0000FF" if judgment == "⭕" else "FF0000",
                                        bold=True, size=14
                                    )
                            elif col_group == target_cols_right and row_idx < len(right_rows):
                                judgment = right_rows[row_idx][3]
                                if c in [7]:
                                    cell.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
                                elif judgment == "⭕":
                                    cell.fill = ok_fill
                                elif judgment == "✖":
                                    cell.fill = ng_fill
                                if c == 9:
                                    cell.font = Font(
                                        color="0000FF" if judgment == "⭕" else "FF0000",
                                        bold=True, size=14
                                    )

            # 列幅
            for col, width in {"A": 8, "B": 8, "C": 8, "D": 8, "F": 8, "G": 8, "H": 8, "I": 8}.items():
                ws.column_dimensions[col].width = width

        output.seek(0)

        import urllib.parse
        filename = f"{user_name}_{exam_title}_採点結果.xlsx"
        encoded_filename = urllib.parse.quote(filename)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)